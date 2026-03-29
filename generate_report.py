#!/usr/bin/env python3
"""
generate_report.py
------------------
/data/ klasöründeki tüm .xlsx dosyalarını okur,
ship verilerini JSON olarak template.html içine gömer
ve index.html çıktısı üretir.

Kullanım:
  python generate_report.py
"""

import os
import json
import re
import glob
from datetime import datetime, date

try:
    from openpyxl import load_workbook
except ImportError:
    raise SystemExit("openpyxl kurulu değil: pip install openpyxl")

# ─── Yardımcılar ──────────────────────────────────────

def cell(ws, row, col):
    """Güvenli hücre okuma."""
    try:
        v = ws.cell(row=row, column=col).value
        return str(v).strip() if v is not None else None
    except Exception:
        return None

def find_in_rows(rows, keyword):
    """rows listesinde keyword içeren satırı bul, yanındaki ilk dolu değeri döndür."""
    kw = keyword.upper()
    for r in rows:
        for i, c in enumerate(r):
            if c and kw in str(c).upper():
                for j in range(i + 1, min(i + 5, len(r))):
                    if r[j] is not None and str(r[j]).strip() != '':
                        return str(r[j]).strip()
    return None

def safe_float(v):
    if v is None:
        return None
    s = str(v).replace(',', '').strip()
    if '=' in s:
        return None
    try:
        return float(s)
    except Exception:
        return None

def fmt_date(v):
    """datetime/date/str → '28.03.2026' formatı."""
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime('%d.%m.%Y')
    s = str(v).strip()
    # already formatted
    if re.match(r'\d{2}\.\d{2}\.\d{4}', s):
        return s[:10]
    return s

def days_left(date_str):
    """'01.04.2026' → float gün kaldı."""
    if not date_str:
        return None
    m = re.search(r'(\d{2})\.(\d{2})\.(\d{4})', str(date_str))
    if not m:
        return None
    try:
        d = date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        delta = (d - date.today()).days
        return round(delta + (d - date.today()).seconds / 86400, 1)
    except Exception:
        return None

PORT_COORDS = {
    'SINGAPORE':    (1.29,   103.85),
    'COLOMBO':      (6.93,   79.85),
    'IMBITUBA':     (-28.23, -48.66),
    'KEMAMAN':      (4.24,   103.43),
    'ROTTERDAM':    (51.90,  4.48),
    'HAMBURG':      (53.54,  9.99),
    'SANTOS':       (-23.95, -46.33),
    'DURBAN':       (-29.85, 31.03),
    'MUMBAI':       (18.96,  72.83),
    'PORT HEDLAND': (-20.31, 118.57),
    'KAOHSIUNG':    (22.62,  120.27),
    'BUSAN':        (35.10,  129.04),
    'QINZHOU':      (21.90,  108.60),
    'CHINA':        (30.5,   121.5),
    'BRAZIL':       (-15.0,  -47.0),
    'MALAYSIA':     (3.14,   101.69),
}

def port_coords(port_str):
    if not port_str:
        return (20.0, 60.0)
    up = port_str.upper()
    for k, v in PORT_COORDS.items():
        if k in up:
            return v
    return (20.0, 60.0)

def parse_lat_lng(lat_str, lng_str):
    """'06 16.3N', '090 55.1E' → (6.272, 90.918)"""
    try:
        lm = re.match(r'(\d+)\s+([\d.]+)\s*([NS]?)', str(lat_str), re.I)
        nm = re.match(r'(\d+)\s+([\d.]+)\s*([EW]?)', str(lng_str), re.I)
        if lm and nm:
            la = int(lm.group(1)) + float(lm.group(2)) / 60
            lo = int(nm.group(1)) + float(nm.group(2)) / 60
            if lm.group(3).upper() == 'S':
                la = -la
            if nm.group(3).upper() == 'W':
                lo = -lo
            return (round(la, 4), round(lo, 4))
    except Exception:
        pass
    return (0.0, 0.0)

# ─── Noon (Seyir) Raporu Parser ───────────────────────

def parse_sea(ws, fname):
    rows = list(ws.iter_rows(values_only=True))

    name      = find_in_rows(rows, 'VESSEL') or fname
    from_port = find_in_rows(rows, 'FROM')   or '—'
    to_port   = find_in_rows(rows, 'TO')     or '—'
    voy       = find_in_rows(rows, 'VOY NO') or '—'
    cargo     = find_in_rows(rows, 'CARGO')  or '—'
    charterer = find_in_rows(rows, 'CHART')  or '—'
    cosp_val  = find_in_rows(rows, 'COSP TIME') or find_in_rows(rows, 'COSP') or '—'

    # Header satırını bul (DAY + Date sütunları)
    hdr_idx = -1
    for i, r in enumerate(rows):
        if r and any(c and str(c).upper() == 'DAY' for c in r) \
              and any(c and 'date' in str(c).lower() for c in r):
            hdr_idx = i
            break

    day_rows = []
    if hdr_idx >= 0:
        for r in rows[hdr_idx + 3: hdr_idx + 30]:
            if not r or not r[0] or '=' in str(r[0]):
                continue
            if not r[1]:
                continue
            day_num = r[0]
            d_date  = fmt_date(r[1])
            lat     = str(r[2]).strip() if r[2] else None
            lng     = str(r[3]).strip() if r[3] else None
            hrs     = safe_float(r[4])
            mins    = safe_float(r[5])
            dist    = safe_float(r[6])
            dtg     = r[7]  # may be formula
            spd     = safe_float(r[8])
            rpm     = safe_float(r[10])
            slip    = safe_float(r[11])
            lsfo    = safe_float(r[12])
            lsdo    = safe_float(r[14])
            rob_fo  = safe_float(r[16])
            rob_do  = safe_float(r[18])
            wind_dir = str(r[23]).strip() if r[23] else ''
            wind_f   = str(r[24]).strip() if r[24] else ''
            wind     = f"{wind_dir} {wind_f}".strip() or '—'
            eta_port = fmt_date(r[27]) if r[27] else None
            remarks  = str(r[40] or r[41] or '').strip()

            cyl_cons = safe_float(r[28])
            cyl_rob  = safe_float(r[29])
            sys_cons = safe_float(r[30])

            if d_date:
                day_rows.append({
                    'date': d_date, 'lat': lat, 'lng': lng,
                    'dist': dist, 'hrs': hrs, 'min': mins,
                    'spd': spd, 'rpm': rpm, 'slip': slip,
                    'lsfo': lsfo, 'lsdo': lsdo,
                    'robLsfo': rob_fo, 'robLsdo': rob_do,
                    'wind': wind, 'etaPort': eta_port,
                    'remarks': remarks,
                    'cylCons': cyl_cons, 'cylROB': cyl_rob,
                    'sysCons': sys_cons,
                })

    last7   = day_rows[-7:]
    latest  = last7[-1] if last7 else {}

    # Fill cylCons from ROB difference if cylCons is None
    for i, d in enumerate(last7):
        if d.get('cylCons') is None and i > 0:
            prev_rob = last7[i-1].get('cylROB')
            curr_rob = d.get('cylROB')
            if prev_rob is not None and curr_rob is not None:
                diff = prev_rob - curr_rob
                if diff >= 0:
                    d['cylCons'] = diff

    # Voyage progress
    dist_made = sum(d['dist'] for d in day_rows if d['dist'])
    dtg_raw   = None
    for r in rows:
        if r and r[0] and str(r[0]).upper().startswith('COSP'):
            raw = r[7]
            if raw and '=' not in str(raw):
                dtg_raw = safe_float(raw)
    total_dist = (dist_made + dtg_raw) if dtg_raw else None
    pct = round(dist_made / total_dist * 100) if total_dist else 0

    lat_dec, lng_dec = (0.0, 0.0)
    if latest.get('lat') and latest.get('lng'):
        lat_dec, lng_dec = parse_lat_lng(latest['lat'], latest['lng'])

    eta_str   = latest.get('etaPort') or '—'
    d_left    = days_left(eta_str)

    return {
        'type':       'sea',
        'subType':    'sea',
        'name':       name,
        'from':       from_port,
        'to':         to_port,
        'voy':        voy,
        'cargo':      cargo,
        'charterer':  charterer,
        'cosp':       cosp_val,
        'eta':        eta_str,
        'daysLeft':   str(d_left) if d_left is not None else '—',
        'distMade':   round(dist_made),
        'totalDist':  round(total_dist) if total_dist else None,
        'pct':        pct,
        'posStr':     f"{latest.get('lat','—')} / {latest.get('lng','—')}",
        'latDec':     lat_dec,
        'lngDec':     lng_dec,
        'speed':      str(latest.get('spd') or '—'),
        'rpm':        str(latest.get('rpm') or '—'),
        'slip':       str(latest.get('slip') or '—'),
        'robLsfo':    latest.get('robLsfo') or 0,
        'robLsdo':    latest.get('robLsdo') or 0,
        'draftF':     '—', 'draftM': '—', 'draftA': '—',
        'last3':      last7,
        'route':      [],
    }

# ─── Port Raporu Parser ───────────────────────────────

def parse_port(ws, fname):
    rows = list(ws.iter_rows(values_only=True))

    name      = find_in_rows(rows, 'VESSEL') or fname
    port      = find_in_rows(rows, 'ARRIVAL PORT') or find_in_rows(rows, 'PORT') or '—'
    voy       = find_in_rows(rows, 'VOY NO') or '—'
    cargo     = find_in_rows(rows, 'CARGO TO BE') or find_in_rows(rows, 'CARGO') or '—'
    charterer = find_in_rows(rows, 'CHART') or '—'

    # Operasyon timestamp'leri
    OP_KEYS = {
        'EOSP': 'EOSP',
        'NOR TEND': 'NOR TENDERED',
        'ANCHR': 'ANCHR',
        'ANCHOR UP': 'ANCHOR UP',
        'POB': 'POB',
        'ALL FAST': 'ALL FAST',
        'BERTH': 'BERTHING',
        'COMM LDNG': 'COMM LDNG',
        'COMP LDNG': 'COMP LDNG',
        'DLOSP': 'DLOSP',
        'COSP': 'COSP',
    }
    ops = {}
    rob_fo = None
    rob_do = None

    for r in rows:
        if not r or not r[0]:
            continue
        k = str(r[0]).upper()
        for key, label in OP_KEYS.items():
            if key in k:
                ops[label] = {
                    'date': fmt_date(r[2]) or '',
                    'time': str(r[3]).strip() if r[3] else '',
                }
        if 'EOSP' in k and rob_fo is None:
            rob_fo = safe_float(r[5])
            rob_do = safe_float(r[7])

    # Daily rows — bölüm 7 veya 8
    daily_rows = []
    dh = -1
    for i, r in enumerate(rows):
        if not r:
            continue
        first = str(r[0] or '').upper()
        if '7.DAILY' in first or '8.DAILY' in first:
            dh = i
            break

    if dh >= 0:
        for r in rows[dh + 2: dh + 40]:
            if not r or not r[1]:
                continue
            d_str = fmt_date(r[1])
            if not d_str or not re.match(r'\d{2}\.\d{2}\.\d{4}', d_str):
                continue
            rl = safe_float(r[10])
            if rl is not None:
                rob_fo = rl
            daily_rows.append({
                'date':    d_str,
                'daily':   safe_float(r[2]),
                'total':   safe_float(r[3]),
                'ets':     str(r[4]).strip() if r[4] else '—',
                'lsfo':    safe_float(r[5]),
                'lsdo':    safe_float(r[7]),
                'robLsfo': rl,
                'robLsdo': safe_float(r[12]),
                'fw':      safe_float(r[9]) or 0,
            })

    last7  = daily_rows[-7:]
    latest = last7[-1] if last7 else {}

    if latest.get('robLsfo'):
        rob_fo = latest['robLsfo']

    # Sub-type
    has_anchr  = bool(ops.get('ANCHR', {}).get('date'))
    has_berth  = bool(ops.get('ALL FAST', {}).get('date') or ops.get('BERTHING', {}).get('date'))
    has_cosp   = bool(ops.get('COSP', {}).get('date'))
    has_anch_up = bool(ops.get('ANCHOR UP', {}).get('date'))

    if has_cosp:
        sub = 'departed'
    elif has_anchr and has_berth:
        # Was anchored, now berthed - show as berthed with anchor history
        sub = 'berthed'
    elif has_anchr and not has_berth:
        sub = 'anchor'
    elif has_berth:
        sub = 'berthed'
    else:
        sub = 'port'

    # Anchor days
    anchor_days = None
    if has_anchr:
        try:
            m = re.search(r'(\d{2})\.(\d{2})\.(\d{4})', ops['ANCHR']['date'])
            if m:
                start = date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
                end_str = ops.get('ANCHOR UP', {}).get('date') or ''
                me = re.search(r'(\d{2})\.(\d{2})\.(\d{4})', end_str)
                end = date(int(me.group(3)), int(me.group(2)), int(me.group(1))) if me else date.today()
                anchor_days = (end - start).days
        except Exception:
            pass

    # Cargo progress
    _cargo_raw = str(find_in_rows(rows, 'CARGO ON BOARD') or '0')
    # Handle "53,900 MT" -> 53900
    import re as _re
    _cargo_num = _re.sub(r'[^\d.]', '', _cargo_raw.split()[0].replace(',','')) if _cargo_raw else '0'
    cargo_total = safe_float(_cargo_num) or None
    cargo_done = latest.get('total')
    cargo_pct  = round(cargo_done / cargo_total * 100) if cargo_total and cargo_done else None

    # ETS days left
    ets_str = latest.get('ets') or '—'
    d_left  = days_left(ets_str)

    lat_dec, lng_dec = port_coords(port)

    # Draft
    draft_f = find_in_rows(rows, 'F') or '—'
    draft_a = find_in_rows(rows, 'A') or '—'

    return {
        'type':       'port',
        'subType':    sub,
        'name':       name,
        'port':       port,
        'voy':        voy,
        'cargo':      cargo,
        'charterer':  charterer,
        'robLsfo':    rob_fo or 0,
        'robLsdo':    rob_do or 0,
        'draftF':     draft_f,
        'draftA':     draft_a,
        'cargoTotal': cargo_total,
        'cargoDone':  cargo_done,
        'cargoPct':   cargo_pct,
        'ets':        ets_str,
        'daysLeft':   str(d_left) if d_left is not None else '—',
        'anchorDays': anchor_days,
        'latDec':     lat_dec,
        'lngDec':     lng_dec,
        'posStr':     port,
        'ops':        ops,
        'last3':      last7,
    }

# ─── Ana İşlem ────────────────────────────────────────

def parse_xlsx(filepath):
    fname = os.path.splitext(os.path.basename(filepath))[0]
    wb = load_workbook(filepath, data_only=True)
    sheets = [s.upper() for s in wb.sheetnames]

    # Sheet seç
    noon_sheet = next((wb.sheetnames[i] for i, s in enumerate(sheets) if 'NOON' in s), None)
    port_sheet = next((wb.sheetnames[i] for i, s in enumerate(sheets) if 'PORT' in s), None)

    if noon_sheet:
        return parse_sea(wb[noon_sheet], fname)
    elif port_sheet:
        return parse_port(wb[port_sheet], fname)
    else:
        # İlk sayfaya bak — sütun adlarından tahmin et
        ws = wb.active
        rows = list(ws.iter_rows(max_row=15, values_only=True))
        flat = ' '.join(str(c) for r in rows for c in r if c)
        if 'STEAMED' in flat.upper() or 'SLIP' in flat.upper():
            return parse_sea(ws, fname)
        else:
            return parse_port(ws, fname)

STALE_HOURS = 36   # Bu saatten eski veri → raporda uyarı göster


def data_age_hours(ship: dict) -> float:
    """
    Raporun içindeki son tarihin kaç saat önce olduğunu döndür.
    Gemi hangi saat diliminde olursa olsun sadece güne bakıyoruz,
    tolerans 36 saat olduğu için saat dilimi farkı sorun yaratmaz.
    """
    latest_date_str = pick_latest_date(ship)
    if latest_date_str == '00000000':
        return 999.0
    try:
        d = datetime.strptime(latest_date_str, '%Y%m%d').replace(hour=12)
        delta = datetime.now() - d
        return round(delta.total_seconds() / 3600, 1)
    except Exception:
        return 999.0


def normalize_vessel_name(name: str) -> str:
    """
    Gemi adını normalleştir → duplicate tespiti için anahtar.
    'MV OCEAN DESTINY', 'Ocean Destiny', 'OCEAN-DESTINY' → 'OCEAN DESTINY'
    """
    s = re.sub(r'^(MV|MT|MS|SS|MY)\s+', '', name.strip(), flags=re.I)
    s = re.sub(r'[-_/]', ' ', s)
    s = re.sub(r'\s+', ' ', s)
    return s.strip().upper()


def pick_latest_date(ship: dict) -> str:
    """
    Rapordaki en son tarihi bul — dosyaların hangisinin daha yeni olduğunu
    dosya adına değil, içindeki veriye göre belirle.
    """
    dates = []
    for d in ship.get('last3', []):
        dt = d.get('date', '')
        m = re.match(r'(\d{2})\.(\d{2})\.(\d{4})', str(dt))
        if m:
            dates.append(f"{m.group(3)}{m.group(2)}{m.group(1)}")  # YYYYMMDD sıralaması için
    return max(dates) if dates else '00000000'


def main():
    data_dir = 'data'
    template_file = 'template.html'
    output_file = 'index.html'

    # Tüm xlsx dosyalarını topla (isim önemli değil)
    xlsx_files = sorted(
        glob.glob(os.path.join(data_dir, '*.xlsx')) +
        glob.glob(os.path.join(data_dir, '*.XLSX')) +
        glob.glob(os.path.join(data_dir, '*.xls'))
    )

    if not xlsx_files:
        print(f"UYARI: {data_dir}/ klasöründe xlsx bulunamadı.")

    # Parse et, dosya adına bakmadan içerikten oku
    raw_ships = []
    for fp in xlsx_files:
        try:
            ship = parse_xlsx(fp)
            ship['_file'] = os.path.basename(fp)
            ship['_mtime'] = os.path.getmtime(fp)          # dosya değişim zamanı (yedek)
            ship['_datadate'] = pick_latest_date(ship)      # rapordaki son tarih (birincil)
            raw_ships.append(ship)
            print(f"  ✓ {os.path.basename(fp)} → {ship['name']} "
                  f"({ship['type']}/{ship.get('subType','')}) "
                  f"[son veri: {ship['_datadate']}]")
        except Exception as e:
            print(f"  ✗ {os.path.basename(fp)} HATA: {e}")

    # ── Duplicate eleme ─────────────────────────────────────────────────────
    # Aynı gemi adı → en yeni raporu (içindeki tarihe göre) tut
    best: dict[str, dict] = {}   # normalize_name → ship_dict
    for s in raw_ships:
        key = normalize_vessel_name(s['name'])
        if key not in best:
            best[key] = s
        else:
            existing = best[key]
            # Önce rapor içi tarihe bak, eşitse dosya değişim zamanına bak
            if (s['_datadate'], s['_mtime']) > (existing['_datadate'], existing['_mtime']):
                print(f"  ↻ '{s['name']}' için daha yeni rapor bulundu: "
                      f"{s['_file']} (eskisi: {existing['_file']})")
                best[key] = s

    ships = []
    for i, s in enumerate(best.values()):
        s.pop('_file', None)
        s.pop('_mtime', None)
        s.pop('_datadate', None)
        s['id'] = i + 1
        age = data_age_hours(s)
        s['_ageHours'] = age
        s['_isStale']  = age > STALE_HOURS
        if s['_isStale']:
            print(f"  ⚠  {s['name']} verisi {age:.0f} saat eski — raporda uyarı gösterilecek")
        ships.append(s)

    print(f"\n  📋 Toplam: {len(raw_ships)} dosya → {len(ships)} benzersiz gemi")

    # Template oku
    if not os.path.exists(template_file):
        raise FileNotFoundError(f"{template_file} bulunamadı!")

    with open(template_file, 'r', encoding='utf-8') as f:
        html = f.read()

    # Rapor tarihi
    today = date.today().strftime('%d.%m.%Y')
    html = html.replace(
        "document.getElementById('hdrSub').textContent =",
        f"// Auto-generated {today} — do not edit\n  document.getElementById('hdrSub').textContent =",
    )

    # Ships JSON'u yerleştir — "Ships will be loaded" yorumunun yerine
    ships_json = json.dumps(ships, ensure_ascii=False, indent=2)
    boot_code = f"""
// ─── AUTO-GENERATED DATA ({today}) ─────────────────────
const BOOT_SHIPS = {ships_json};
BOOT_SHIPS.forEach((s, i) => {{
  s.color = COLORS[i % COLORS.length];
  addShip(s);
}});
"""
    html = html.replace(
        '// Ships will be loaded from Excel files. Page starts empty.',
        boot_code
    )

    # Rapor tarihi badge'ini de güncelle
    html = re.sub(
        r'class="dbadge">[^<]*</div>',
        f'class="dbadge">{today}</div>',
        html
    )

    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(html)

    count = len(ships)
    print(f"\n✅ index.html oluşturuldu — {count} gemi — {today}")

if __name__ == '__main__':
    main()
